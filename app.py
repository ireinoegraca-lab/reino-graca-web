import os, base64, io, zlib, struct
from flask import Flask, render_template, request, jsonify, session, redirect, url_for, Response
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from models import db, Perfil, Config, Usuario, Membro, Ministerio, Evento, Musica, Setlist, SetlistItem, Financeiro, Campanha, CampanhaCotista, MuralPost, PedidoOracao, Escala, EscalaItem, Presenca, OrcamentoItem, Comunicado, ComunicadoLeitura, Celula, CelulaMembro
from datetime import datetime, date

def make_icon_png(size):
    """Gera PNG simples: fundo vermelho + cruz branca centralizada."""
    bar = max(1, size // 10)
    arm = size // 3
    rows = []
    for y in range(size):
        row = bytearray()
        for x in range(size):
            cx, cy = x - size // 2, y - size // 2
            cross = (abs(cx) < bar and abs(cy) < arm) or (abs(cy) < bar and abs(cx) < arm)
            row += b'\xff\xff\xff' if cross else b'\xe1\x1d\x2a'
        rows.append(bytes(row))
    raw = b''.join(b'\x00' + r for r in rows)
    def chunk(tag, data):
        body = tag + data
        return struct.pack('>I', len(data)) + body + struct.pack('>I', zlib.crc32(body) & 0xffffffff)
    ihdr = struct.pack('>IIBBBBB', size, size, 8, 2, 0, 0, 0)
    return b'\x89PNG\r\n\x1a\n' + chunk(b'IHDR', ihdr) + chunk(b'IDAT', zlib.compress(raw)) + chunk(b'IEND', b'')

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'rg-secret-2026')
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///reino.db').replace('postgres://', 'postgresql://')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'index'

@login_manager.user_loader
def load_user(uid): return Usuario.query.get(int(uid))

def has_perm(p):
    if not current_user.is_authenticated: return False
    if current_user.status != 'ativo': return False
    perms = current_user.get_perms()
    return bool(perms.get(p, False))

def is_admin(): return has_perm('p_usuarios')
def can_manage(mid): return is_admin() or (current_user.is_authenticated and current_user.ministerio_id==mid and has_perm('p_ministerios'))

MESES_CONFIRMACAO = 6

def precisa_confirmar(u):
    """True se o usuário nunca confirmou ou passaram mais de 6 meses."""
    if u.role in ('admin',): return False          # admins não precisam confirmar
    if not u.ultima_confirmacao: return True
    from datetime import date, timedelta
    try:
        ultima = date.fromisoformat(u.ultima_confirmacao)
        return (date.today() - ultima).days > MESES_CONFIRMACAO * 30
    except: return True

def serialize_usuario(u):
    return {'id':u.id,'nome':u.nome,'usuario':u.usuario,'role':u.role,'status':u.status or 'ativo',
            'perfil_id':u.perfil_id,'perfil_nome':u.perfil.nome if u.perfil else u.role,
            'ministerio_id':u.ministerio_id,'perms':u.get_perms(),
            'ultima_confirmacao':u.ultima_confirmacao,'precisa_confirmar':precisa_confirmar(u)}
def serialize_membro(m):
    return {'id':m.id,'nome':m.nome,'nasc':m.nasc or '','tel':m.tel or '','email':m.email or '',
            'profissao':m.profissao or '','status':m.status or 'Ativo','bairro':m.bairro or '',
            'obs':m.obs or '','foto':m.foto or '','ministerio_id':m.ministerio_id,
            'ministerio_nome':m.ministerio.nome if m.ministerio else ''}
def serialize_ministerio(c):
    return {'id':c.id,'nome':c.nome,'descricao':c.descricao or '','cor':c.cor or '#e11d2a',
            'membros':[{'id':mb.id,'nome':mb.nome,'status':mb.status} for mb in c.membros]}
def serialize_evento(e):
    return {'id':e.id,'titulo':e.titulo,'data':e.data,'hora':e.hora or '','local':e.local or '','tipo':e.tipo or 'culto','descricao':e.descricao or ''}
def serialize_musica(m):
    return {'id':m.id,'titulo':m.titulo,'artista':m.artista or '','tom':m.tom or '',
            'cifra':m.cifra or '','letra':m.letra or '','bpm':m.bpm or 0,
            'link_youtube':m.link_youtube or '','categoria':m.categoria or '',
            'ministerio_id':m.ministerio_id,'ministerio_nome':m.ministerio.nome if m.ministerio else ''}
def serialize_setlist(s):
    itens = sorted(s.itens, key=lambda x: x.ordem)
    return {'id':s.id,'titulo':s.titulo,'data':s.data or '','hora':s.hora or '',
            'musicas':[{'id':i.musica_id,'titulo':i.musica.titulo,'artista':i.musica.artista or '','tom':i.tom or i.musica.tom or ''} for i in itens]}
def serialize_fin(f):
    return {'id':f.id,'tipo':f.tipo,'categoria':f.categoria or '','valor':f.valor,'data':f.data,
            'descricao':f.descricao or '','membro_id':f.membro_id,'forma':f.forma or '',
            'campanha_id':f.campanha_id,'campanha_nome':f.campanha.nome if f.campanha else '',
            'membro_nome':f.membro.nome if f.membro else ''}

def serialize_campanha(c):
    arrecadado = db.session.query(db.func.sum(Financeiro.valor)).filter_by(campanha_id=c.id, tipo='entrada').scalar() or 0
    mensal = sum(co.valor_mensal for co in c.cotistas)
    return {'id':c.id,'nome':c.nome,'descricao':c.descricao or '','meta':c.meta or 0,
            'data_inicio':c.data_inicio or '','data_fim':c.data_fim or '','status':c.status or 'ativa',
            'dia_vencimento':c.dia_vencimento,'arrecadado':float(arrecadado),'total_mensal':mensal,
            'cotistas':[{'id':co.id,'membro_id':co.membro_id,'nome':co.membro.nome,
                         'tel':co.membro.tel or '','valor_mensal':co.valor_mensal} for co in c.cotistas]}
def serialize_post(p):
    return {'id':p.id,'titulo':p.titulo,'texto':p.texto or '','imagem':p.imagem or '',
            'ministerio_id':p.ministerio_id,'ministerio_nome':p.ministerio.nome if p.ministerio else 'Geral',
            'autor_nome':p.autor.nome,'criado_em':p.criado_em.isoformat()}

# ── PWA ───────────────────────────────────────────────────────────
@app.route('/manifest.json')
def manifest():
    cfg = {r.chave: r.valor for r in Config.query.all()}
    nome = cfg.get('nome_igreja', 'Igreja')
    data = {
        "name": f"{nome} — Sistema",
        "short_name": nome[:12],
        "description": f"Sistema de gestão da {nome}",
        "start_url": "/", "display": "standalone",
        "background_color": "#080808", "theme_color": "#e11d2a",
        "icons": [
            {"src": "/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any maskable"},
            {"src": "/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any maskable"}
        ]
    }
    import json
    return Response(json.dumps(data), mimetype='application/manifest+json')

@app.route('/icon-<int:size>.png')
def icon_png(size):
    size = min(size, 512)
    return Response(make_icon_png(size), mimetype='image/png',
                    headers={'Cache-Control': 'public, max-age=86400'})

# ── MAIN ──────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/cadastro')
def cadastro_publico():
    cfg = {r.chave: r.valor for r in Config.query.all()}
    return render_template('cadastro.html',
                           nome_igreja=cfg.get('nome_igreja','Reino & Graça'),
                           nome_pastor=cfg.get('nome_pastor','Pastor'))

@app.route('/membro/<int:mid>')
def perfil_publico(mid):
    cfg = {r.chave: r.valor for r in Config.query.all()}
    nome_igreja = cfg.get('nome_igreja','Reino & Graça')
    m = Membro.query.get(mid)
    membro = None
    if m:
        ministerio_nome = m.ministerio.nome if m.ministerio else ''
        u_vinc = Usuario.query.filter_by(nome=m.nome).first()
        role_labels = {'membro':'Membro','lider':'Líder','secretaria':'Secretaria',
                       'tesoureiro':'Tesoureiro','pastor':'Pastor','admin':'Administrador'}
        funcao = role_labels.get(u_vinc.role, u_vinc.role or 'Membro') if u_vinc else 'Membro'
        if not ministerio_nome and u_vinc and u_vinc.ministerio:
            ministerio_nome = u_vinc.ministerio.nome
        membro = {'id':m.id,'nome':m.nome,'foto':m.foto or '','status':m.status,
                  'profissao':m.profissao or '','bairro':m.bairro or '',
                  'nasc':m.nasc or '','ministerio':ministerio_nome,'funcao':funcao}
    return render_template('membro_perfil.html', membro=membro,
                           nome_igreja=nome_igreja,
                           data_hoje=date.today().strftime('%d/%m/%Y'))

@app.route('/escanear')
@login_required
def escanear():
    if not has_perm('p_escanear'):
        return redirect('/')
    cfg = {r.chave: r.valor for r in Config.query.all()}
    hoje = date.today().isoformat()
    eventos = Evento.query.filter(Evento.data >= hoje).order_by(Evento.data).limit(20).all()
    # inclui também eventos recentes (últimos 7 dias)
    from datetime import timedelta
    semana = (date.today() - timedelta(days=7)).isoformat()
    recentes = Evento.query.filter(Evento.data >= semana, Evento.data < hoje).order_by(Evento.data.desc()).limit(10).all()
    todos = recentes + eventos
    return render_template('escanear.html',
                           nome_igreja=cfg.get('nome_igreja','Reino & Graça'),
                           eventos=[{'id':e.id,'titulo':e.titulo,'data':e.data} for e in todos])

@app.route('/api/presenca/qr', methods=['POST'])
@login_required
def presenca_qr():
    if not has_perm('p_escanear'):
        return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    d = request.json or {}
    membro_id = d.get('membro_id')
    evento_id = d.get('evento_id')
    if not membro_id or not evento_id:
        return jsonify({'ok':False,'msg':'Dados incompletos'}), 400
    m = Membro.query.get(membro_id)
    if not m:
        return jsonify({'ok':False,'msg':'Membro não encontrado'}), 404
    role_labels = {'membro':'Membro','lider':'Líder','secretaria':'Secretaria',
                   'tesoureiro':'Tesoureiro','pastor':'Pastor','admin':'Administrador'}
    # busca usuario vinculado ao membro pelo email ou nome
    u_vinc = Usuario.query.filter_by(nome=m.nome).first()
    funcao = role_labels.get(u_vinc.role, u_vinc.role or 'Membro') if u_vinc else 'Membro'
    min_nome = (m.ministerio.nome if m.ministerio else '') or \
               (u_vinc.ministerio.nome if u_vinc and u_vinc.ministerio else '')
    existente = Presenca.query.filter_by(evento_id=evento_id, membro_id=membro_id).first()
    if existente:
        return jsonify({'ok':True,'duplicado':True,'nome':m.nome,'funcao':funcao,'ministerio':min_nome})
    db.session.add(Presenca(evento_id=evento_id, membro_id=membro_id, presente=True))
    db.session.commit()
    return jsonify({'ok':True,'duplicado':False,'nome':m.nome,'funcao':funcao,'ministerio':min_nome})

# ── AUTH ──────────────────────────────────────────────────────────
@app.route('/api/login', methods=['POST'])
def api_login():
    d = request.json
    u = Usuario.query.filter_by(usuario=d.get('usuario','').lower()).first()
    if not u or not u.check_senha(d.get('senha','')):
        return jsonify({'ok':False,'msg':'Usuário ou senha incorretos'}), 401
    if u.status == 'bloqueado':
        return jsonify({'ok':False,'msg':'Conta bloqueada. Fale com o administrador.'}), 403
    if u.status == 'pendente':
        return jsonify({'ok':False,'msg':'Cadastro aguardando aprovação do administrador.'}), 403
    login_user(u, remember=True)
    return jsonify({'ok':True,'user':{'id':u.id,'nome':u.nome,'role':u.role,'status':u.status,
                                      'ministerio_id':u.ministerio_id,'perfil_id':u.perfil_id,
                                      'perms':u.get_perms(),'precisa_confirmar':precisa_confirmar(u)}})

@app.route('/api/register', methods=['POST'])
def api_register():
    d = request.json or {}
    nome    = (d.get('nome') or '').strip()
    email   = (d.get('email') or '').strip().lower()
    tel     = (d.get('tel') or '').strip()
    nasc    = (d.get('nasc') or '').strip()
    usuario = (d.get('usuario') or '').strip().lower()
    senha   = d.get('senha') or ''
    if not nome or not usuario or not senha:
        return jsonify({'ok':False,'msg':'Nome, usuário e senha são obrigatórios'}), 400
    if Usuario.query.filter_by(usuario=usuario).first():
        return jsonify({'ok':False,'msg':'Usuário já existe. Escolha outro nome de usuário.'}), 409
    profissao = (d.get('profissao') or '').strip()
    bairro    = (d.get('bairro')    or '').strip()
    perfil_membro = Perfil.query.filter_by(nome='Membro').first()
    hoje = date.today().isoformat()
    u = Usuario(nome=nome, usuario=usuario, role='membro', status='pendente',
                perfil_id=perfil_membro.id if perfil_membro else None, ultima_confirmacao=hoje)
    u.set_senha(senha)
    db.session.add(u)
    foto = (d.get('foto') or '').strip() or None
    m = Membro(nome=nome, email=email, tel=tel, nasc=nasc or None, status='Ativo',
               profissao=profissao or None, bairro=bairro or None, foto=foto)
    db.session.add(m)
    db.session.commit()
    return jsonify({'ok':True,'pendente':True,'msg':'Cadastro enviado! Aguarde a aprovação do administrador.'})

@app.route('/api/logout', methods=['POST'])
@login_required
def api_logout():
    logout_user()
    return jsonify({'ok':True})

@app.route('/api/carteirinha/me')
@login_required
def carteirinha_me():
    u = current_user
    # tenta vincular pelo ministerio_id do usuário
    m = None
    if u.ministerio_id:
        m = Membro.query.filter_by(ministerio_id=u.ministerio_id).filter(
            Membro.nome.ilike(f'%{u.nome.split()[0]}%')).first()
    # busca por nome exato ou parcial
    if not m:
        m = Membro.query.filter(Membro.nome.ilike(u.nome)).first()
    if not m:
        primeiro = u.nome.split()[0]
        m = Membro.query.filter(Membro.nome.ilike(f'{primeiro}%')).first()
    cfg = {r.chave: r.valor for r in Config.query.all()}
    role_labels = {'membro':'Membro','lider':'Líder','secretaria':'Secretaria',
                   'tesoureiro':'Tesoureiro','pastor':'Pastor','admin':'Administrador'}
    funcao = role_labels.get(u.role, u.role or 'Membro')
    min_nome = (m.ministerio.nome if m and m.ministerio else '') or \
               (u.ministerio.nome if u.ministerio else '')
    if m:
        return jsonify({'ok': True, 'membro': {
            'id': m.id, 'nome': m.nome, 'foto': m.foto or '',
            'status': m.status or 'Ativo',
            'profissao': m.profissao or '',
            'ministerio_nome': min_nome,
            'funcao_igreja': funcao,
            'nasc': m.nasc or '', 'bairro': m.bairro or ''
        }, 'config': cfg})
    # fallback: usa dados do usuário
    return jsonify({'ok': True, 'membro': {
        'id': u.id, 'nome': u.nome, 'foto': '',
        'status': 'Membro', 'profissao': '',
        'ministerio_nome': min_nome,
        'funcao_igreja': funcao,
        'nasc': '', 'bairro': ''
    }, 'config': cfg})

@app.route('/api/meu-perfil', methods=['PUT'])
@login_required
def update_meu_perfil():
    d = request.json or {}
    # atualiza senha se enviada
    nova_senha = d.get('nova_senha','').strip()
    if nova_senha:
        if len(nova_senha) < 6:
            return jsonify({'ok':False,'msg':'Senha deve ter pelo menos 6 caracteres'}), 400
        current_user.set_senha(nova_senha)
    # encontra o membro vinculado
    u = current_user
    m = None
    if u.ministerio_id:
        m = Membro.query.filter_by(ministerio_id=u.ministerio_id).filter(
            Membro.nome.ilike(f'%{u.nome.split()[0]}%')).first()
    if not m:
        m = Membro.query.filter(Membro.nome.ilike(u.nome)).first()
    if not m:
        primeiro = u.nome.split()[0]
        m = Membro.query.filter(Membro.nome.ilike(f'{primeiro}%')).first()
    if m:
        if 'tel'       in d: m.tel       = d['tel'] or None
        if 'email'     in d: m.email     = d['email'] or None
        if 'profissao' in d: m.profissao = d['profissao'] or None
        if 'bairro'    in d: m.bairro    = d['bairro'] or None
        if 'obs'       in d: m.obs       = d['obs'] or None
        if 'foto'      in d: m.foto      = d['foto'] or None
    db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/me')
def api_me():
    if not current_user.is_authenticated:
        return jsonify({'ok':False}), 401
    u = current_user
    return jsonify({'ok':True,'user':{'id':u.id,'nome':u.nome,'role':u.role,'status':u.status,
                                      'ministerio_id':u.ministerio_id,'perfil_id':u.perfil_id,
                                      'perms':u.get_perms(),'precisa_confirmar':precisa_confirmar(u)}})

# ── MEMBROS ───────────────────────────────────────────────────────
@app.route('/api/membros')
@login_required
def get_membros():
    return jsonify([serialize_membro(m) for m in Membro.query.order_by(Membro.nome).all()])

@app.route('/api/membros', methods=['POST'])
@login_required
def add_membro():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    m = Membro(nome=d['nome'],nasc=d.get('nasc'),tel=d.get('tel'),email=d.get('email'),
               profissao=d.get('profissao'),status=d.get('status','Ativo'),
               bairro=d.get('bairro'),obs=d.get('obs'),ministerio_id=d.get('ministerio_id'))
    db.session.add(m); db.session.commit()
    return jsonify({'ok':True,'membro':serialize_membro(m)})

@app.route('/api/membros/<int:mid>', methods=['PUT'])
@login_required
def update_membro(mid):
    if not is_admin(): return jsonify({'ok':False}), 403
    m = Membro.query.get_or_404(mid); d = request.json
    for k in ['nome','nasc','tel','email','profissao','status','bairro','obs','foto','ministerio_id']:
        if k in d: setattr(m, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'membro':serialize_membro(m)})

@app.route('/api/membros/<int:mid>', methods=['DELETE'])
@login_required
def del_membro(mid):
    if not is_admin(): return jsonify({'ok':False}), 403
    m = Membro.query.get_or_404(mid); db.session.delete(m); db.session.commit()
    return jsonify({'ok':True})

# ── MINISTÉRIOS ───────────────────────────────────────────────────
@app.route('/api/ministerios')
@login_required
def get_ministerios():
    return jsonify([serialize_ministerio(c) for c in Ministerio.query.order_by(Ministerio.nome).all()])

@app.route('/api/ministerios', methods=['POST'])
@login_required
def add_ministerio():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    c = Ministerio(nome=d['nome'],descricao=d.get('descricao',''),cor=d.get('cor','#e11d2a'))
    db.session.add(c); db.session.commit()
    return jsonify({'ok':True,'ministerio':serialize_ministerio(c)})

@app.route('/api/ministerios/<int:cid>', methods=['PUT'])
@login_required
def update_ministerio(cid):
    if not can_manage(cid): return jsonify({'ok':False}), 403
    c = Ministerio.query.get_or_404(cid); d = request.json
    for k in ['nome','descricao','cor']:
        if k in d: setattr(c, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'ministerio':serialize_ministerio(c)})

@app.route('/api/ministerios/<int:cid>', methods=['DELETE'])
@login_required
def del_ministerio(cid):
    if not is_admin(): return jsonify({'ok':False}), 403
    c = Ministerio.query.get_or_404(cid); db.session.delete(c); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/ministerios/<int:cid>/membros', methods=['POST'])
@login_required
def add_membro_ministerio(cid):
    if not can_manage(cid): return jsonify({'ok':False}), 403
    d = request.json
    m = Membro.query.get_or_404(d['membro_id'])
    m.ministerio_id = cid; db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/ministerios/<int:cid>/membros/<int:mid>', methods=['DELETE'])
@login_required
def rem_membro_ministerio(cid, mid):
    if not can_manage(cid): return jsonify({'ok':False}), 403
    m = Membro.query.get_or_404(mid)
    if m.ministerio_id == cid: m.ministerio_id = None; db.session.commit()
    return jsonify({'ok':True})

# ── AGENDA ────────────────────────────────────────────────────────
@app.route('/api/eventos')
@login_required
def get_eventos():
    return jsonify([serialize_evento(e) for e in Evento.query.order_by(Evento.data).all()])

@app.route('/api/eventos', methods=['POST'])
@login_required
def add_evento():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    e = Evento(titulo=d['titulo'],data=d['data'],hora=d.get('hora'),local=d.get('local'),tipo=d.get('tipo','culto'),descricao=d.get('descricao'))
    db.session.add(e); db.session.commit()
    return jsonify({'ok':True,'evento':serialize_evento(e)})

@app.route('/api/eventos/<int:eid>', methods=['PUT'])
@login_required
def update_evento(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    e = Evento.query.get_or_404(eid); d = request.json
    for k in ['titulo','data','hora','local','tipo','descricao']:
        if k in d: setattr(e, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'evento':serialize_evento(e)})

@app.route('/api/eventos/<int:eid>', methods=['DELETE'])
@login_required
def del_evento(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    e = Evento.query.get_or_404(eid); db.session.delete(e); db.session.commit()
    return jsonify({'ok':True})

# ── LOUVOR ────────────────────────────────────────────────────────
@app.route('/api/musicas')
@login_required
def get_musicas():
    return jsonify([serialize_musica(m) for m in Musica.query.order_by(Musica.titulo).all()])

@app.route('/api/musicas', methods=['POST'])
@login_required
def add_musica():
    d = request.json
    m = Musica(titulo=d['titulo'],artista=d.get('artista'),tom=d.get('tom'),cifra=d.get('cifra'),
               letra=d.get('letra'),bpm=d.get('bpm') or None,link_youtube=d.get('link_youtube'),
               categoria=d.get('categoria'),ministerio_id=d.get('ministerio_id') or None)
    db.session.add(m); db.session.commit()
    return jsonify({'ok':True,'musica':serialize_musica(m)})

@app.route('/api/musicas/<int:mid>', methods=['PUT'])
@login_required
def update_musica(mid):
    m = Musica.query.get_or_404(mid); d = request.json
    for k in ['titulo','artista','tom','cifra','letra','link_youtube','categoria']:
        if k in d: setattr(m, k, d[k] or None)
    if 'bpm' in d: m.bpm = int(d['bpm']) if d['bpm'] else None
    if 'ministerio_id' in d: m.ministerio_id = d['ministerio_id'] or None
    db.session.commit()
    return jsonify({'ok':True,'musica':serialize_musica(m)})

@app.route('/api/musicas/<int:mid>', methods=['DELETE'])
@login_required
def del_musica(mid):
    m = Musica.query.get_or_404(mid); db.session.delete(m); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/setlists')
@login_required
def get_setlists():
    return jsonify([serialize_setlist(s) for s in Setlist.query.order_by(Setlist.data).all()])

@app.route('/api/setlists', methods=['POST'])
@login_required
def add_setlist():
    d = request.json
    s = Setlist(titulo=d['titulo'],data=d.get('data'),hora=d.get('hora'))
    db.session.add(s); db.session.flush()
    for i, item in enumerate(d.get('musicas',[])):
        si = SetlistItem(setlist_id=s.id,musica_id=item['id'],ordem=i,tom=item.get('tom',''))
        db.session.add(si)
    db.session.commit()
    return jsonify({'ok':True,'setlist':serialize_setlist(s)})

@app.route('/api/setlists/<int:sid>', methods=['DELETE'])
@login_required
def del_setlist(sid):
    s = Setlist.query.get_or_404(sid); db.session.delete(s); db.session.commit()
    return jsonify({'ok':True})

# ── FINANCEIRO ────────────────────────────────────────────────────
@app.route('/api/financeiro')
@login_required
def get_financeiro():
    if not is_admin(): return jsonify({'ok':False}), 403
    return jsonify([serialize_fin(f) for f in Financeiro.query.order_by(Financeiro.data.desc()).all()])

@app.route('/api/financeiro', methods=['POST'])
@login_required
def add_financeiro():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    f = Financeiro(tipo=d['tipo'],categoria=d.get('categoria'),valor=float(d['valor']),
                   data=d['data'],descricao=d.get('descricao'),membro_id=d.get('membro_id'),
                   forma=d.get('forma'),campanha_id=d.get('campanha_id') or None)
    db.session.add(f); db.session.commit()
    return jsonify({'ok':True,'lancamento':serialize_fin(f)})

@app.route('/api/financeiro/importar', methods=['POST'])
@login_required
def importar_financeiro():
    if not is_admin(): return jsonify({'ok':False}), 403
    if 'arquivo' not in request.files:
        return jsonify({'ok':False,'msg':'Nenhum arquivo enviado'}), 400
    arquivo = request.files['arquivo']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
        def col(row, names):
            for n in names:
                if n in headers:
                    v = row[headers.index(n)].value
                    return str(v).strip() if v is not None else ''
            return ''
        importados = 0
        for row in ws.iter_rows(min_row=2):
            valor_raw = col(row, ['valor','value','quantia'])
            data_raw  = col(row, ['data','date','data do lançamento'])
            if not valor_raw or not data_raw: continue
            try: valor = float(str(valor_raw).replace('R$','').replace('.','').replace(',','.').strip())
            except: continue
            tipo = col(row, ['tipo','type']) or 'entrada'
            if tipo.lower() not in ('entrada','saida','saída'): tipo='entrada'
            if tipo.lower() in ('saida','saída'): tipo='saida'
            f = Financeiro(
                tipo=tipo,
                categoria=col(row, ['categoria','category','tipo de entrada']),
                valor=valor,
                data=data_raw[:10],
                descricao=col(row, ['descrição','descricao','description','obs','observação']),
                forma=col(row, ['forma','forma de pagamento','payment','método']) or 'Dinheiro',
            )
            db.session.add(f)
            importados += 1
        db.session.commit()
        return jsonify({'ok':True,'importados':importados})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)}), 500

@app.route('/api/financeiro/<int:fid>', methods=['PUT'])
@login_required
def update_financeiro(fid):
    if not is_admin(): return jsonify({'ok':False}), 403
    f = Financeiro.query.get_or_404(fid); d = request.json
    for k in ['tipo','categoria','valor','data','descricao','forma']:
        if k in d: setattr(f, k, float(d[k]) if k=='valor' else d[k])
    if 'campanha_id' in d: f.campanha_id = d['campanha_id'] or None
    if 'membro_id' in d: f.membro_id = d['membro_id'] or None
    db.session.commit()
    return jsonify({'ok':True,'lancamento':serialize_fin(f)})

@app.route('/api/financeiro/<int:fid>', methods=['DELETE'])
@login_required
def del_financeiro(fid):
    if not is_admin(): return jsonify({'ok':False}), 403
    f = Financeiro.query.get_or_404(fid); db.session.delete(f); db.session.commit()
    return jsonify({'ok':True})

# ── CAMPANHAS ────────────────────────────────────────────────────
@app.route('/api/campanhas')
@login_required
def get_campanhas():
    if not has_perm('p_financeiro'): return jsonify({'ok':False}), 403
    return jsonify([serialize_campanha(c) for c in Campanha.query.order_by(Campanha.id.desc()).all()])

@app.route('/api/campanhas', methods=['POST'])
@login_required
def add_campanha():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    dv = d.get('dia_vencimento')
    c = Campanha(nome=d['nome'],descricao=d.get('descricao'),meta=float(d.get('meta') or 0),
                 data_inicio=d.get('data_inicio'),data_fim=d.get('data_fim'),status='ativa',
                 dia_vencimento=int(dv) if dv else None)
    db.session.add(c); db.session.commit()
    return jsonify({'ok':True,'campanha':serialize_campanha(c)})

@app.route('/api/campanhas/<int:cid>', methods=['PUT'])
@login_required
def update_campanha(cid):
    if not is_admin(): return jsonify({'ok':False}), 403
    c = Campanha.query.get_or_404(cid); d = request.json
    for k in ['nome','descricao','data_inicio','data_fim','status']:
        if k in d: setattr(c, k, d[k])
    if 'meta' in d: c.meta = float(d['meta'] or 0)
    if 'dia_vencimento' in d: c.dia_vencimento = int(d['dia_vencimento']) if d['dia_vencimento'] else None
    db.session.commit()
    return jsonify({'ok':True,'campanha':serialize_campanha(c)})

@app.route('/api/campanhas/<int:cid>', methods=['DELETE'])
@login_required
def del_campanha(cid):
    if not is_admin(): return jsonify({'ok':False}), 403
    c = Campanha.query.get_or_404(cid); db.session.delete(c); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/campanhas/lembretes')
@login_required
def get_lembretes():
    """Retorna cotistas com vencimento nos próximos N dias ou atrasados."""
    if not has_perm('p_financeiro'): return jsonify({'ok':False}), 403
    from datetime import date as dt
    hoje = dt.today()
    dias_aviso = int(request.args.get('dias', 5))
    resultado = []
    for c in Campanha.query.filter_by(status='ativa').all():
        if not c.dia_vencimento: continue
        try:
            import calendar
            ultimo_dia = calendar.monthrange(hoje.year, hoje.month)[1]
            dia = min(c.dia_vencimento, ultimo_dia)
            venc = dt(hoje.year, hoje.month, dia)
        except: continue
        diff = (venc - hoje).days
        if -7 <= diff <= dias_aviso:
            for co in c.cotistas:
                resultado.append({
                    'campanha_id': c.id, 'campanha_nome': c.nome,
                    'membro_id': co.membro_id, 'membro_nome': co.membro.nome,
                    'tel': co.membro.tel or '', 'valor_mensal': co.valor_mensal,
                    'vencimento': venc.isoformat(), 'dias_para_vencer': diff,
                    'status': 'vencido' if diff < 0 else 'hoje' if diff == 0 else 'proximo'
                })
    return jsonify(resultado)

@app.route('/api/campanhas/<int:cid>/cotistas', methods=['POST'])
@login_required
def add_cotista(cid):
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    if CampanhaCotista.query.filter_by(campanha_id=cid, membro_id=d['membro_id']).first():
        return jsonify({'ok':False,'msg':'Membro já é cotista desta campanha'}), 400
    co = CampanhaCotista(campanha_id=cid, membro_id=d['membro_id'], valor_mensal=float(d['valor_mensal']))
    db.session.add(co); db.session.commit()
    c = Campanha.query.get(cid)
    return jsonify({'ok':True,'campanha':serialize_campanha(c)})

@app.route('/api/campanhas/<int:cid>/cotistas/<int:coid>', methods=['DELETE'])
@login_required
def del_cotista(cid, coid):
    if not is_admin(): return jsonify({'ok':False}), 403
    co = CampanhaCotista.query.get_or_404(coid)
    db.session.delete(co); db.session.commit()
    c = Campanha.query.get(cid)
    return jsonify({'ok':True,'campanha':serialize_campanha(c)})

# ── ORAÇÃO ────────────────────────────────────────────────────────
def serialize_oracao(p):
    return {'id':p.id,'texto':p.texto,'nome_solicitante':p.nome_solicitante or 'Anônimo',
            'privado':p.privado,'status':p.status,
            'criado_em':p.criado_em.strftime('%Y-%m-%d'),'membro_id':p.membro_id}

@app.route('/api/oracoes')
@login_required
def get_oracoes():
    q = PedidoOracao.query
    if not is_admin():
        q = q.filter((PedidoOracao.privado==False) | (PedidoOracao.membro_id==None))
    return jsonify([serialize_oracao(p) for p in q.order_by(PedidoOracao.criado_em.desc()).all()])

@app.route('/api/oracoes', methods=['POST'])
@login_required
def add_oracao():
    d = request.json or {}
    texto = (d.get('texto') or '').strip()
    if not texto: return jsonify({'ok':False,'msg':'Escreva o pedido'}), 400
    anonimo = bool(d.get('anonimo'))
    p = PedidoOracao(
        texto=texto,
        nome_solicitante=None if anonimo else current_user.nome,
        membro_id=None if anonimo else None,
        privado=bool(d.get('privado')),
        status='aberto'
    )
    db.session.add(p); db.session.commit()
    return jsonify({'ok':True,'pedido':serialize_oracao(p)})

@app.route('/api/oracoes/<int:pid>/status', methods=['PUT'])
@login_required
def update_oracao_status(pid):
    if not is_admin(): return jsonify({'ok':False}), 403
    p = PedidoOracao.query.get_or_404(pid)
    p.status = request.json.get('status', p.status)
    db.session.commit()
    return jsonify({'ok':True,'pedido':serialize_oracao(p)})

@app.route('/api/oracoes/<int:pid>', methods=['DELETE'])
@login_required
def del_oracao(pid):
    if not is_admin(): return jsonify({'ok':False}), 403
    p = PedidoOracao.query.get_or_404(pid); db.session.delete(p); db.session.commit()
    return jsonify({'ok':True})

# ── MURAL ─────────────────────────────────────────────────────────
@app.route('/api/mural')
@login_required
def get_mural():
    return jsonify([serialize_post(p) for p in MuralPost.query.order_by(MuralPost.criado_em.desc()).all()])

@app.route('/api/mural', methods=['POST'])
@login_required
def add_post():
    d = request.json
    mid = d.get('ministerio_id')
    if mid and not can_manage(int(mid)): return jsonify({'ok':False}), 403
    p = MuralPost(titulo=d['titulo'],texto=d.get('texto'),imagem=d.get('imagem'),
                  ministerio_id=mid,autor_id=current_user.id)
    db.session.add(p); db.session.commit()
    return jsonify({'ok':True,'post':serialize_post(p)})

@app.route('/api/mural/<int:pid>', methods=['PUT'])
@login_required
def update_post(pid):
    p = MuralPost.query.get_or_404(pid)
    if not is_admin() and p.autor_id != current_user.id: return jsonify({'ok':False}), 403
    d = request.json
    for k in ['titulo','texto','imagem','ministerio_id']:
        if k in d: setattr(p, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'post':serialize_post(p)})

@app.route('/api/mural/<int:pid>', methods=['DELETE'])
@login_required
def del_post(pid):
    p = MuralPost.query.get_or_404(pid)
    if not is_admin() and p.autor_id != current_user.id: return jsonify({'ok':False}), 403
    db.session.delete(p); db.session.commit()
    return jsonify({'ok':True})

# ── PERFIS ────────────────────────────────────────────────────────
@app.route('/api/perfis')
@login_required
def get_perfis():
    return jsonify([p.to_dict() for p in Perfil.query.order_by(Perfil.id).all()])

@app.route('/api/perfis', methods=['POST'])
@login_required
def add_perfil():
    if not has_perm('p_perfis'): return jsonify({'ok':False}), 403
    d = request.json
    p = Perfil(nome=d['nome'], cor=d.get('cor','#e11d2a'),
               p_membros=d.get('p_membros',False), p_ministerios=d.get('p_ministerios',False),
               p_agenda=d.get('p_agenda',True), p_louvor=d.get('p_louvor',False),
               p_mural=d.get('p_mural',True), p_financeiro=d.get('p_financeiro',False),
               p_usuarios=d.get('p_usuarios',False), p_perfis=d.get('p_perfis',False),
               p_escanear=d.get('p_escanear',False),
               pode_aprovar=d.get('pode_aprovar',False))
    db.session.add(p); db.session.commit()
    return jsonify({'ok':True,'perfil':p.to_dict()})

@app.route('/api/perfis/<int:pid>', methods=['PUT'])
@login_required
def update_perfil(pid):
    if not has_perm('p_perfis'): return jsonify({'ok':False}), 403
    p = Perfil.query.get_or_404(pid)
    if p.builtin: return jsonify({'ok':False,'msg':'Perfis padrão não podem ser editados'}), 400
    d = request.json
    for k in ['nome','cor','p_membros','p_ministerios','p_agenda','p_louvor','p_mural','p_financeiro','p_usuarios','p_perfis','pode_aprovar','p_escanear']:
        if k in d: setattr(p, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'perfil':p.to_dict()})

@app.route('/api/perfis/<int:pid>', methods=['DELETE'])
@login_required
def del_perfil(pid):
    if not has_perm('p_perfis'): return jsonify({'ok':False}), 403
    p = Perfil.query.get_or_404(pid)
    if p.builtin: return jsonify({'ok':False,'msg':'Perfis padrão não podem ser excluídos'}), 400
    db.session.delete(p); db.session.commit()
    return jsonify({'ok':True})

# ── USUÁRIOS ──────────────────────────────────────────────────────
@app.route('/api/usuarios')
@login_required
def get_usuarios():
    if not is_admin(): return jsonify({'ok':False}), 403
    return jsonify([serialize_usuario(u) for u in Usuario.query.order_by(Usuario.nome).all()])

@app.route('/api/confirmar-membro', methods=['POST'])
@login_required
def confirmar_membro():
    current_user.ultima_confirmacao = date.today().isoformat()
    db.session.commit()
    return jsonify({'ok':True,'data':current_user.ultima_confirmacao})

@app.route('/api/usuarios/sem-confirmacao')
@login_required
def get_sem_confirmacao():
    if not is_admin(): return jsonify({'ok':False}), 403
    lista = []
    for u in Usuario.query.filter(Usuario.status=='ativo', Usuario.role!='admin').all():
        if precisa_confirmar(u):
            lista.append({**serialize_usuario(u), 'dias_sem_confirmar':
                (date.today() - date.fromisoformat(u.ultima_confirmacao)).days if u.ultima_confirmacao else None})
    return jsonify(lista)

@app.route('/api/usuarios/pendentes')
@login_required
def get_pendentes():
    if not (is_admin() or has_perm('pode_aprovar')): return jsonify({'ok':False}), 403
    return jsonify([serialize_usuario(u) for u in Usuario.query.filter_by(status='pendente').all()])

@app.route('/api/usuarios', methods=['POST'])
@login_required
def add_usuario():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    if Usuario.query.filter_by(usuario=d['usuario'].lower()).first():
        return jsonify({'ok':False,'msg':'Usuário já existe'}), 400
    u = Usuario(nome=d['nome'], usuario=d['usuario'].lower(), role=d.get('role','membro'),
                status='ativo', perfil_id=d.get('perfil_id'), ministerio_id=d.get('ministerio_id'))
    u.set_senha(d['senha'])
    db.session.add(u); db.session.commit()
    return jsonify({'ok':True,'usuario':serialize_usuario(u)})

@app.route('/api/usuarios/<int:uid>', methods=['PUT'])
@login_required
def update_usuario(uid):
    if not is_admin(): return jsonify({'ok':False}), 403
    u = Usuario.query.get_or_404(uid); d = request.json
    if 'nome' in d: u.nome = d['nome']
    if 'role' in d: u.role = d['role']
    if 'status' in d: u.status = d['status']
    if 'perfil_id' in d: u.perfil_id = d['perfil_id'] or None
    if 'ministerio_id' in d: u.ministerio_id = d['ministerio_id']
    if 'senha' in d and d['senha']: u.set_senha(d['senha'])
    db.session.commit()
    return jsonify({'ok':True,'usuario':serialize_usuario(u)})

@app.route('/api/usuarios/<int:uid>/aprovar', methods=['POST'])
@login_required
def aprovar_usuario(uid):
    if not (is_admin() or has_perm('pode_aprovar')): return jsonify({'ok':False}), 403
    u = Usuario.query.get_or_404(uid)
    u.status = 'ativo'
    if not u.perfil_id:
        perfil = Perfil.query.filter_by(nome='Membro').first()
        if perfil: u.perfil_id = perfil.id
    db.session.commit()
    return jsonify({'ok':True,'usuario':serialize_usuario(u)})

@app.route('/api/usuarios/<int:uid>/bloquear', methods=['POST'])
@login_required
def bloquear_usuario(uid):
    if not is_admin(): return jsonify({'ok':False}), 403
    if uid == current_user.id: return jsonify({'ok':False,'msg':'Não pode bloquear sua própria conta'}), 400
    u = Usuario.query.get_or_404(uid)
    u.status = 'bloqueado' if u.status == 'ativo' else 'ativo'
    db.session.commit()
    return jsonify({'ok':True,'status':u.status})

@app.route('/api/usuarios/<int:uid>', methods=['DELETE'])
@login_required
def del_usuario(uid):
    if not is_admin(): return jsonify({'ok':False}), 403
    if uid == current_user.id: return jsonify({'ok':False,'msg':'Não pode excluir sua própria conta'}), 400
    u = Usuario.query.get_or_404(uid); db.session.delete(u); db.session.commit()
    return jsonify({'ok':True})

# ── CONFIGURAÇÕES ─────────────────────────────────────────────────
@app.route('/api/config')
@login_required
def get_config():
    rows = Config.query.all()
    return jsonify({r.chave: r.valor for r in rows})

@app.route('/api/config', methods=['PUT'])
@login_required
def save_config():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json or {}
    for chave, valor in d.items():
        c = Config.query.filter_by(chave=chave).first()
        if c: c.valor = valor
        else: db.session.add(Config(chave=chave, valor=valor))
    db.session.commit()
    return jsonify({'ok':True})

# ── IMPORTAR MEMBROS (Excel) ──────────────────────────────────────
@app.route('/api/membros/importar', methods=['POST'])
@login_required
def importar_membros():
    if not is_admin(): return jsonify({'ok':False}), 403
    if 'arquivo' not in request.files:
        return jsonify({'ok':False,'msg':'Nenhum arquivo enviado'}), 400
    arquivo = request.files['arquivo']
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(arquivo.read()))
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]
        def col(row, names):
            for n in names:
                if n in headers:
                    v = row[headers.index(n)].value
                    return str(v).strip() if v is not None else ''
            return ''
        importados = 0
        for row in ws.iter_rows(min_row=2):
            nome = col(row, ['nome','name'])
            if not nome: continue
            m = Membro(
                nome=nome,
                nasc=col(row, ['nascimento','data de nascimento','nasc','data_nasc','birthday']),
                tel=col(row, ['telefone','tel','phone','celular','whatsapp']),
                email=col(row, ['email','e-mail']),
                profissao=col(row, ['profissão','profissao','profession','ocupação','ocupacao']),
                bairro=col(row, ['bairro','neighborhood','endereco','endereço']),
                status=col(row, ['status']) or 'Ativo',
                obs=col(row, ['observações','observacoes','obs','notas','notes'])
            )
            db.session.add(m)
            importados += 1
        db.session.commit()
        return jsonify({'ok':True,'importados':importados})
    except Exception as e:
        return jsonify({'ok':False,'msg':str(e)}), 500

# ── ESCALAS ───────────────────────────────────────────────────────
def serialize_escala(e):
    return {'id':e.id,'titulo':e.titulo,'data':e.data,'descricao':e.descricao or '',
            'itens':[{'id':i.id,'membro_id':i.membro_id,'nome':i.membro.nome,
                      'tel':i.membro.tel or '','funcao':i.funcao} for i in e.itens]}

@app.route('/api/escalas')
@login_required
def get_escalas():
    return jsonify([serialize_escala(e) for e in Escala.query.order_by(Escala.data.desc()).all()])

@app.route('/api/escalas', methods=['POST'])
@login_required
def add_escala():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    e = Escala(titulo=d['titulo'], data=d['data'], descricao=d.get('descricao',''))
    db.session.add(e); db.session.commit()
    return jsonify({'ok':True,'escala':serialize_escala(e)})

@app.route('/api/escalas/<int:eid>', methods=['PUT'])
@login_required
def update_escala(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    e = Escala.query.get_or_404(eid); d = request.json
    for k in ['titulo','data','descricao']:
        if k in d: setattr(e, k, d[k])
    db.session.commit()
    return jsonify({'ok':True,'escala':serialize_escala(e)})

@app.route('/api/escalas/<int:eid>', methods=['DELETE'])
@login_required
def del_escala(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    e = Escala.query.get_or_404(eid); db.session.delete(e); db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/escalas/<int:eid>/itens', methods=['POST'])
@login_required
def add_escala_item(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    item = EscalaItem(escala_id=eid, membro_id=d['membro_id'], funcao=d['funcao'])
    db.session.add(item); db.session.commit()
    return jsonify({'ok':True,'escala':serialize_escala(Escala.query.get(eid))})

@app.route('/api/escalas/<int:eid>/itens/<int:iid>', methods=['DELETE'])
@login_required
def del_escala_item(eid, iid):
    if not is_admin(): return jsonify({'ok':False}), 403
    item = EscalaItem.query.get_or_404(iid); db.session.delete(item); db.session.commit()
    return jsonify({'ok':True,'escala':serialize_escala(Escala.query.get(eid))})

# ── PRESENÇA ──────────────────────────────────────────────────────
@app.route('/api/eventos/<int:eid>/presenca')
@login_required
def get_presenca(eid):
    rows = Presenca.query.filter_by(evento_id=eid).all()
    return jsonify([{'membro_id':p.membro_id,'presente':p.presente} for p in rows])

@app.route('/api/eventos/<int:eid>/presenca', methods=['POST'])
@login_required
def set_presenca(eid):
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json  # [{membro_id, presente}, ...]
    Presenca.query.filter_by(evento_id=eid).delete()
    for row in d:
        if row.get('presente'):
            db.session.add(Presenca(evento_id=eid, membro_id=row['membro_id'], presente=True))
    db.session.commit()
    return jsonify({'ok':True,'total':len([r for r in d if r.get('presente')])})

# ── ORÇAMENTO ─────────────────────────────────────────────────────
@app.route('/api/orcamento')
@login_required
def get_orcamento():
    if not has_perm('p_financeiro'): return jsonify({'ok':False}), 403
    ano = int(request.args.get('ano', date.today().year))
    items = OrcamentoItem.query.filter_by(ano=ano).all()
    fin = Financeiro.query.filter(Financeiro.data.like(f'{ano}-%'), Financeiro.tipo=='saida').all()
    realizado = {}
    for f in fin:
        cat = f.categoria or 'Outros'
        realizado[cat] = realizado.get(cat, 0) + f.valor
    return jsonify({'ano':ano,
        'itens':[{'id':i.id,'categoria':i.categoria,'previsto':i.valor_previsto,
                  'realizado':realizado.get(i.categoria,0)} for i in items],
        'categorias_sem_orcamento':list(set(realizado.keys()) - {i.categoria for i in items})})

@app.route('/api/orcamento', methods=['POST'])
@login_required
def add_orcamento():
    if not is_admin(): return jsonify({'ok':False}), 403
    d = request.json
    ano = int(d.get('ano', date.today().year))
    existing = OrcamentoItem.query.filter_by(ano=ano, categoria=d['categoria']).first()
    if existing:
        existing.valor_previsto = float(d['valor_previsto'])
    else:
        db.session.add(OrcamentoItem(ano=ano, categoria=d['categoria'], valor_previsto=float(d['valor_previsto'])))
    db.session.commit()
    return jsonify({'ok':True})

@app.route('/api/orcamento/<int:oid>', methods=['DELETE'])
@login_required
def del_orcamento(oid):
    if not is_admin(): return jsonify({'ok':False}), 403
    i = OrcamentoItem.query.get_or_404(oid); db.session.delete(i); db.session.commit()
    return jsonify({'ok':True})

# ── RECIBO DE DÍZIMO ──────────────────────────────────────────────
@app.route('/api/recibo')
@login_required
def get_recibo():
    ano = int(request.args.get('ano', date.today().year))
    mid = request.args.get('membro_id')
    if mid and not is_admin(): return jsonify({'ok':False}), 403
    q = Financeiro.query.filter(Financeiro.data.like(f'{ano}-%'), Financeiro.tipo=='entrada',
                                 Financeiro.categoria=='Dízimo')
    if mid: q = q.filter_by(membro_id=int(mid))
    lancamentos = q.order_by(Financeiro.data).all()
    membros = {m.id:m.nome for m in Membro.query.all()}
    cfg = {r.chave: r.valor for r in Config.query.all()}
    return jsonify({
        'ano':ano,'total':sum(f.valor for f in lancamentos),
        'lancamentos':[{**serialize_fin(f),'membro_nome_r':membros.get(f.membro_id,'')} for f in lancamentos],
        'igreja':cfg.get('nome_igreja','Igreja'),'pastor':cfg.get('nome_pastor','Pastor')
    })

# ── PORTAL PÚBLICO ────────────────────────────────────────────────
@app.route('/portal')
def portal():
    hoje = date.today().isoformat()
    mes = hoje[5:7]
    membros = Membro.query.order_by(Membro.nome).all()
    musicas = Musica.query.order_by(Musica.titulo).all()
    posts = MuralPost.query.order_by(MuralPost.criado_em.desc()).limit(20).all()
    dados = {
        'codigoAcesso': os.environ.get('PORTAL_CODIGO', 'reino2026'),
        'pinLouvor': os.environ.get('PORTAL_PIN', 'louvor'),
        'onesignalAppId': 'e6cb61e9-ccfd-47a7-aa96-6eff5776b122',
        'mural': [{'id':p.id,'titulo':p.titulo,'texto':p.texto or '','imagem':p.imagem or '',
                   'ministerio':p.ministerio.nome if p.ministerio else 'Geral','autor':p.autor.nome} for p in posts],
        'eventos': [serialize_evento(e) for e in Evento.query.filter(Evento.data >= hoje).order_by(Evento.data).limit(20).all()],
        'membros': [{'nome':m.nome,'nasc':m.nasc or ''} for m in membros if m.nasc],
        'musicas': [{'titulo':m.titulo,'artista':m.artista or '','tom':m.tom or '','cifra':m.cifra or ''} for m in musicas],
        'guia': [],
    }
    return render_template('portal.html', dados=dados)

# ── INIT DB ───────────────────────────────────────────────────────
PERFIS_PADRAO = [
    {'nome':'Pastor',     'cor':'#c9922a','builtin':True,
     'p_membros':True,'p_ministerios':True,'p_agenda':True,'p_louvor':True,'p_mural':True,
     'p_financeiro':True,'p_usuarios':True,'p_perfis':True,'pode_aprovar':True,'p_escanear':True},
    {'nome':'Administrador','cor':'#e11d2a','builtin':True,
     'p_membros':True,'p_ministerios':True,'p_agenda':True,'p_louvor':True,'p_mural':True,
     'p_financeiro':True,'p_usuarios':True,'p_perfis':True,'pode_aprovar':True,'p_escanear':True},
    {'nome':'Secretaria', 'cor':'#60a5fa','builtin':True,
     'p_membros':True,'p_ministerios':True,'p_agenda':True,'p_louvor':False,'p_mural':True,
     'p_financeiro':False,'p_usuarios':False,'p_perfis':False,'pode_aprovar':True,'p_escanear':True},
    {'nome':'Líder',      'cor':'#a78bfa','builtin':True,
     'p_membros':True,'p_ministerios':True,'p_agenda':True,'p_louvor':True,'p_mural':True,
     'p_financeiro':False,'p_usuarios':False,'p_perfis':False,'pode_aprovar':False,'p_escanear':True},
    {'nome':'Tesoureiro', 'cor':'#22c55e','builtin':True,
     'p_membros':False,'p_ministerios':False,'p_agenda':True,'p_louvor':False,'p_mural':True,
     'p_financeiro':True,'p_usuarios':False,'p_perfis':False,'pode_aprovar':False,'p_escanear':False},
    {'nome':'Membro',     'cor':'#9b9296','builtin':True,
     'p_membros':False,'p_ministerios':False,'p_agenda':True,'p_louvor':True,'p_mural':True,
     'p_financeiro':False,'p_usuarios':False,'p_perfis':False,'pode_aprovar':False,'p_escanear':False},
]

def migrate_columns():
    """Adiciona colunas novas em tabelas existentes sem destruir dados."""
    with app.app_context():
        conn = db.engine.connect()
        def add_col(table, col, definition):
            try:
                conn.execute(db.text(f'ALTER TABLE {table} ADD COLUMN {col} {definition}'))
                conn.commit()
                print(f'Coluna {table}.{col} adicionada.')
            except Exception:
                conn.rollback()
        add_col('usuarios',   'status',      "VARCHAR(20) DEFAULT 'ativo'")
        add_col('usuarios',   'perfil_id',   'INTEGER REFERENCES perfis(id)')
        add_col('membros',    'foto',        'TEXT')
        add_col('financeiro', 'campanha_id',      'INTEGER REFERENCES campanhas(id)')
        add_col('campanhas',  'dia_vencimento',       'INTEGER')
        add_col('usuarios',   'ultima_confirmacao',   'VARCHAR(10)')
        add_col('comunicados','ministerio_id', 'INTEGER REFERENCES ministerios(id)')
        add_col('musicas','letra',         'TEXT')
        add_col('musicas','bpm',           'INTEGER')
        add_col('musicas','link_youtube',  'VARCHAR(300)')
        add_col('musicas','categoria',     'VARCHAR(50)')
        add_col('musicas','ministerio_id', 'INTEGER REFERENCES ministerios(id)')
        add_col('perfis','p_escanear', 'BOOLEAN DEFAULT FALSE')
        conn.close()

def init_db():
    with app.app_context():
        db.create_all()
        # Migra colunas novas em tabelas existentes
        migrate_columns()
        # Seed / atualiza perfis padrão
        for pd in PERFIS_PADRAO:
            p = Perfil.query.filter_by(nome=pd['nome']).first()
            if not p:
                db.session.add(Perfil(**pd))
            else:
                # sempre sincroniza p_escanear nos perfis builtin
                p.p_escanear = pd.get('p_escanear', False)
        db.session.commit()
        # Seed config padrão
        configs_padrao = [
            ('whatsapp_secretaria', ''),
            ('nome_pastor', 'Pastor'),
            ('nome_igreja', 'Reino & Graça'),
        ]
        for chave, valor in configs_padrao:
            if not Config.query.filter_by(chave=chave).first():
                db.session.add(Config(chave=chave, valor=valor))
        db.session.commit()
        # Seed usuários iniciais
        if not Usuario.query.first():
            p_admin = Perfil.query.filter_by(nome='Administrador').first()
            admin = Usuario(nome='Administrador', usuario='admin', role='admin', status='ativo',
                            perfil_id=p_admin.id if p_admin else None)
            admin.set_senha('admin123')
            db.session.add(admin)
            dirceu = Usuario(nome='Dirceu Gonçalves', usuario='dirceu', role='admin', status='ativo',
                             perfil_id=p_admin.id if p_admin else None)
            dirceu.set_senha('dirceu123')
            db.session.add(dirceu)
            db.session.commit()
            print('Usuários iniciais criados: admin/admin123 e dirceu/dirceu123')

# ── COMUNICADOS ───────────────────────────────────────────────────
def serialize_comunicado(c, user_id):
    lidos_ids = {l.usuario_id for l in c.leituras}
    return {
        'id': c.id, 'titulo': c.titulo, 'texto': c.texto,
        'autor': c.autor.nome, 'ministerio': c.ministerio.nome if c.ministerio else None,
        'criado_em': c.criado_em.strftime('%Y-%m-%d %H:%M'),
        'total_leituras': len(c.leituras),
        'lido_por_mim': user_id in lidos_ids
    }

@app.route('/api/comunicados', methods=['GET'])
@login_required
def get_comunicados():
    cs = Comunicado.query.order_by(Comunicado.criado_em.desc()).all()
    return jsonify([serialize_comunicado(c, current_user.id) for c in cs])

@app.route('/api/comunicados', methods=['POST'])
@login_required
def criar_comunicado():
    if not (is_admin() or has_perm('p_mural')): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    d = request.json
    c = Comunicado(titulo=d['titulo'], texto=d['texto'], autor_id=current_user.id,
                   ministerio_id=d.get('ministerio_id') or None)
    db.session.add(c); db.session.commit()
    return jsonify({'ok': True, 'id': c.id})

@app.route('/api/comunicados/<int:cid>', methods=['DELETE'])
@login_required
def del_comunicado(cid):
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    c = Comunicado.query.get_or_404(cid)
    db.session.delete(c); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/comunicados/<int:cid>/ler', methods=['POST'])
@login_required
def marcar_lido(cid):
    existe = ComunicadoLeitura.query.filter_by(comunicado_id=cid, usuario_id=current_user.id).first()
    if not existe:
        db.session.add(ComunicadoLeitura(comunicado_id=cid, usuario_id=current_user.id))
        db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/comunicados/<int:cid>/leituras')
@login_required
def get_leituras(cid):
    c = Comunicado.query.get_or_404(cid)
    if not (is_admin() or c.autor_id == current_user.id):
        return jsonify({'ok': False, 'msg': 'Sem permissão'}), 403
    leituras = (ComunicadoLeitura.query
                .filter_by(comunicado_id=cid)
                .join(Usuario, ComunicadoLeitura.usuario_id == Usuario.id)
                .add_columns(Usuario.nome, ComunicadoLeitura.lido_em)
                .order_by(ComunicadoLeitura.lido_em.asc())
                .all())
    total_usuarios = Usuario.query.filter_by(status='ativo').count()
    return jsonify({
        'ok': True,
        'total_usuarios': total_usuarios,
        'leituras': [{'nome': l.nome, 'lido_em': l.lido_em.strftime('%d/%m/%Y %H:%M')} for l in leituras]
    })

# ── CÉLULAS ───────────────────────────────────────────────────────
def serialize_celula(c):
    return {
        'id': c.id, 'nome': c.nome, 'descricao': c.descricao,
        'lider': c.lider.nome if c.lider else None,
        'lider_id': c.lider_id,
        'dia_semana': c.dia_semana, 'horario': c.horario, 'local': c.local,
        'membros': [{'id': m.membro_id, 'nome': m.membro.nome, 'foto': m.membro.foto} for m in c.membros]
    }

@app.route('/api/celulas', methods=['GET'])
@login_required
def get_celulas():
    cs = Celula.query.order_by(Celula.nome).all()
    return jsonify({'celulas': [serialize_celula(c) for c in cs]})

@app.route('/api/celulas', methods=['POST'])
@login_required
def criar_celula():
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    d = request.json
    c = Celula(nome=d['nome'], descricao=d.get('descricao'), lider_id=d.get('lider_id') or None,
               dia_semana=d.get('dia_semana'), horario=d.get('horario'), local=d.get('local'))
    db.session.add(c); db.session.commit()
    return jsonify({'ok': True, 'id': c.id})

@app.route('/api/celulas/<int:cid>', methods=['PUT'])
@login_required
def edit_celula(cid):
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    c = Celula.query.get_or_404(cid); d = request.json
    c.nome=d.get('nome',c.nome); c.descricao=d.get('descricao',c.descricao)
    c.lider_id=d.get('lider_id') or None; c.dia_semana=d.get('dia_semana',c.dia_semana)
    c.horario=d.get('horario',c.horario); c.local=d.get('local',c.local)
    db.session.commit(); return jsonify({'ok': True})

@app.route('/api/celulas/<int:cid>', methods=['DELETE'])
@login_required
def del_celula(cid):
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    c = Celula.query.get_or_404(cid); db.session.delete(c); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/celulas/<int:cid>/membros', methods=['POST'])
@login_required
def add_celula_membro(cid):
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    d = request.json; mid = d['membro_id']
    if not CelulaMembro.query.filter_by(celula_id=cid, membro_id=mid).first():
        db.session.add(CelulaMembro(celula_id=cid, membro_id=mid)); db.session.commit()
    return jsonify({'ok': True})

@app.route('/api/celulas/<int:cid>/membros/<int:mid>', methods=['DELETE'])
@login_required
def rem_celula_membro(cid, mid):
    if not is_admin(): return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    m = CelulaMembro.query.filter_by(celula_id=cid, membro_id=mid).first()
    if m: db.session.delete(m); db.session.commit()
    return jsonify({'ok': True})

# ── PIX ───────────────────────────────────────────────────────────
@app.route('/api/pix-info', methods=['GET'])
@login_required
def get_pix_info():
    chave = Config.query.filter_by(chave='pix_chave').first()
    nome  = Config.query.filter_by(chave='nome_igreja').first()
    return jsonify({'chave': chave.valor if chave else '', 'nome': nome.valor if nome else 'Igreja'})

# ── FIX PERMISSÕES ────────────────────────────────────────────────
@app.route('/api/admin/fix-perms')
@login_required
def fix_perms():
    if not is_admin(): return jsonify({'ok':False}),403
    for pd in PERFIS_PADRAO:
        p = Perfil.query.filter_by(nome=pd['nome']).first()
        if p:
            p.p_escanear = pd.get('p_escanear', False)
    db.session.commit()
    return jsonify({'ok':True,'msg':'Permissões dos perfis builtin atualizadas'})

# ── BACKUP / EXPORT ───────────────────────────────────────────────
@app.route('/api/admin/export-db')
@login_required
def export_db():
    if not is_admin():
        return jsonify({'ok':False,'msg':'Sem permissão'}), 403
    from datetime import datetime as dt
    data = {
        'exported_at': dt.utcnow().isoformat(),
        'version': '1.0',
        'membros': [serialize_membro(m) for m in Membro.query.all()],
        'ministerios': [{'id':c.id,'nome':c.nome,'descricao':c.descricao or '','cor':c.cor or '#e11d2a'} for c in Ministerio.query.all()],
        'usuarios': [{'id':u.id,'nome':u.nome,'usuario':u.usuario,'role':u.role,'status':u.status,
                      'ministerio_id':u.ministerio_id,'ultima_confirmacao':u.ultima_confirmacao or ''} for u in Usuario.query.all()],
        'eventos': [{'id':e.id,'titulo':e.titulo,'data':e.data,'hora':e.hora or '','local':e.local or '',
                     'tipo':e.tipo or '','descricao':e.descricao or ''} for e in Evento.query.all()],
        'presencas': [{'id':p.id,'evento_id':p.evento_id,'membro_id':p.membro_id,'presente':p.presente} for p in Presenca.query.all()],
        'musicas': [serialize_musica(m) for m in Musica.query.all()],
        'setlists': [serialize_setlist(s) for s in Setlist.query.all()],
        'financeiro': [{'id':f.id,'tipo':f.tipo,'categoria':f.categoria or '','valor':f.valor,
                        'data':f.data,'descricao':f.descricao or '','membro_id':f.membro_id,
                        'forma':f.forma or '','campanha_id':f.campanha_id} for f in Financeiro.query.all()],
        'campanhas': [serialize_campanha(c) for c in Campanha.query.all()],
        'escalas': [{'id':e.id,'titulo':e.titulo,'data':e.data,'descricao':e.descricao or '',
                     'itens':[{'membro_id':i.membro_id,'funcao':i.funcao} for i in e.itens]} for e in Escala.query.all()],
        'mural': [{'id':p.id,'titulo':p.titulo,'texto':p.texto or '','imagem':p.imagem or '',
                   'ministerio_id':p.ministerio_id,'autor_id':p.autor_id,'criado_em':p.criado_em.isoformat()} for p in MuralPost.query.all()],
        'config': {r.chave: r.valor for r in Config.query.all()},
        'perfis': [p.to_dict() for p in Perfil.query.filter_by(builtin=False).all()],
        'celulas': [{'id':c.id,'nome':c.nome,'descricao':c.descricao or '','lider_id':c.lider_id,
                     'dia_semana':c.dia_semana or '','horario':c.horario or '','local':c.local or '',
                     'membros':[m.membro_id for m in c.membros]} for c in Celula.query.all()],
        'comunicados': [{'id':c.id,'titulo':c.titulo,'texto':c.texto,'autor_id':c.autor_id,
                         'ministerio_id':c.ministerio_id,'criado_em':c.criado_em.isoformat()} for c in Comunicado.query.all()],
        'pedidos_oracao': [{'id':p.id,'texto':p.texto,'nome_solicitante':p.nome_solicitante or '',
                            'membro_id':p.membro_id,'privado':p.privado,'status':p.status,
                            'criado_em':p.criado_em.isoformat()} for p in PedidoOracao.query.all()],
    }
    import json
    from flask import Response
    filename = f"reino-graca-backup-{dt.utcnow().strftime('%Y%m%d-%H%M')}.json"
    return Response(json.dumps(data, ensure_ascii=False, indent=2),
                    mimetype='application/json',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})

init_db()

if __name__ == '__main__':
    app.run(debug=True, port=5700)
